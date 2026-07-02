"""YouTube comment, reply, and live-chat extraction helpers."""

from __future__ import annotations

import csv
import re
import time
from datetime import datetime
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

from googleapiclient.discovery import build

try:
    from googleapiclient.errors import HttpError
except Exception:  # pragma: no cover - googleapiclient normally provides this
    HttpError = Exception


YOUTUBE_COMMENTS_PER_PAGE = 100
YOUTUBE_REPLIES_PER_PAGE = 100
YOUTUBE_LIVE_CHAT_PER_PAGE = 200
API_DELAY_BETWEEN_PAGES = 0.01  # keeps the run fast without hammering repeated page requests


class CommentsDisabledError(Exception):
    pass


class VideoNotFoundError(Exception):
    pass


class QuotaExceededError(Exception):
    pass


class YouTubeCommentExtractor:
    def __init__(
        self,
        api_key: str,
        spam_threshold: float = 0.5,
        blacklist_patterns: Optional[List[str]] = None,
        whitelist_patterns: Optional[List[str]] = None,
    ):
        self.youtube = build("youtube", "v3", developerKey=api_key)
        self.spam_threshold = spam_threshold
        self.blacklist_patterns = [p for p in (blacklist_patterns or []) if p]
        self.whitelist_patterns = [p for p in (whitelist_patterns or []) if p]
        self._should_stop = False

    def stop(self) -> None:
        self._should_stop = True

    @staticmethod
    def extract_video_id(url: str) -> str:
        patterns = [
            r"(?:v=)([0-9A-Za-z_-]{11})",
            r"youtu\.be/([0-9A-Za-z_-]{11})",
            r"embed/([0-9A-Za-z_-]{11})",
            r"shorts/([0-9A-Za-z_-]{11})",
            r"/live/([0-9A-Za-z_-]{11})",
            r"^([0-9A-Za-z_-]{11})$",
        ]
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        return ""

    def process_video(
        self,
        video_url: str,
        max_results: Optional[int] = None,
        progress_callback: Optional[Callable[[int], None]] = None,
        filter_spam: bool = True,
        min_likes: int = 0,
        sort_by: str = "relevance",
        exclude_creator: bool = False,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        filter_words: Optional[List[str]] = None,
        cancel_event: Any = None,
        extract_comments: bool = True,
        extract_live_chat: bool = False,
    ) -> Tuple[Dict[str, Any], List[Dict[str, Any]], List[Dict[str, Any]]]:
        """Return (metadata, clean_items, spam_items), matching what main.py expects."""
        video_id = self.extract_video_id(video_url)
        if not video_id:
            raise ValueError(f"Invalid YouTube URL: {video_url}")

        metadata = self.fetch_video_metadata(video_id)
        raw_items: List[Dict[str, Any]] = []

        if extract_comments:
            raw_items.extend(
                self.fetch_comments_and_replies(
                    video_id=video_id,
                    metadata=metadata,
                    sort_by=sort_by,
                    max_results=max_results,
                    min_likes=min_likes,
                    exclude_creator=exclude_creator,
                    date_from=date_from,
                    date_to=date_to,
                    filter_words=filter_words,
                    cancel_event=cancel_event,
                    progress_callback=progress_callback,
                )
            )

        if extract_live_chat and not self._cancelled(cancel_event):
            raw_items.extend(
                self.fetch_live_chat_messages(
                    metadata=metadata,
                    max_results=None if max_results is None else max(0, max_results - len(raw_items)),
                    date_from=date_from,
                    date_to=date_to,
                    filter_words=filter_words,
                    cancel_event=cancel_event,
                )
            )

        clean_items, spam_items = self.apply_local_filters(raw_items, filter_spam=filter_spam)
        self.sort_items(clean_items, sort_by)
        self.sort_items(spam_items, sort_by)

        metadata["items_fetched"] = len(raw_items)
        metadata["clean_items"] = len(clean_items)
        metadata["spam_items"] = len(spam_items)
        metadata["extract_comments"] = extract_comments
        metadata["extract_live_chat"] = extract_live_chat
        return metadata, clean_items, spam_items

    def fetch_video_metadata(self, video_id: str) -> Dict[str, Any]:
        try:
            response = self.youtube.videos().list(
                part="snippet,statistics,liveStreamingDetails",
                id=video_id,
            ).execute()
        except Exception as e:
            self._raise_friendly_error(e)

        items = response.get("items", [])
        if not items:
            raise VideoNotFoundError("The requested video could not be found.")

        item = items[0]
        snippet = item.get("snippet", {})
        statistics = item.get("statistics", {})
        live_details = item.get("liveStreamingDetails", {})

        return {
            "video_id": video_id,
            "title": snippet.get("title", ""),
            "channel_title": snippet.get("channelTitle", ""),
            "channel_id": snippet.get("channelId", ""),
            "published_at": snippet.get("publishedAt", ""),
            "description": snippet.get("description", ""),
            "comment_count": statistics.get("commentCount", ""),
            "view_count": statistics.get("viewCount", ""),
            "like_count": statistics.get("likeCount", ""),
            "active_live_chat_id": live_details.get("activeLiveChatId", ""),
            "actual_start_time": live_details.get("actualStartTime", ""),
            "actual_end_time": live_details.get("actualEndTime", ""),
        }

    def fetch_comments_and_replies(
        self,
        video_id: str,
        metadata: Dict[str, Any],
        sort_by: str,
        max_results: Optional[int],
        min_likes: int,
        exclude_creator: bool,
        date_from: Optional[str],
        date_to: Optional[str],
        filter_words: Optional[List[str]],
        cancel_event: Any,
        progress_callback: Optional[Callable[[int], None]],
    ) -> List[Dict[str, Any]]:
        comments: List[Dict[str, Any]] = []
        request = self.youtube.commentThreads().list(
            part="snippet",
            videoId=video_id,
            maxResults=YOUTUBE_COMMENTS_PER_PAGE,
            textFormat="plainText",
            order=self.api_comment_order(sort_by),
        )

        while request and not self._cancelled(cancel_event):
            try:
                response = request.execute()
            except Exception as e:
                self._raise_friendly_error(e)

            for item in response.get("items", []):
                if self._cancelled(cancel_event) or self._limit_reached(comments, max_results):
                    break

                thread_snippet = item.get("snippet", {})
                top_comment = thread_snippet.get("topLevelComment", {})
                top_snippet = top_comment.get("snippet", {})
                comment_id = top_comment.get("id") or item.get("id", "")
                reply_count = int(thread_snippet.get("totalReplyCount", 0) or 0)

                parent = self._build_comment_row(
                    video_id=video_id,
                    metadata=metadata,
                    comment_id=comment_id,
                    snippet=top_snippet,
                    item_type="Parent Comment",
                    parent_id="",
                    reply_count=reply_count,
                )

                if self._passes_fetch_filters(
                    parent, min_likes, exclude_creator, date_from, date_to, filter_words, metadata
                ):
                    comments.append(parent)
                    if progress_callback:
                        progress_callback(len(comments))

                if reply_count > 0 and not self._limit_reached(comments, max_results):
                    comments.extend(
                        self.fetch_all_replies(
                            parent_id=comment_id,
                            video_id=video_id,
                            metadata=metadata,
                            max_results=None if max_results is None else max_results - len(comments),
                            min_likes=min_likes,
                            exclude_creator=exclude_creator,
                            date_from=date_from,
                            date_to=date_to,
                            filter_words=filter_words,
                            cancel_event=cancel_event,
                        )
                    )

            request = None if self._limit_reached(comments, max_results) else self.youtube.commentThreads().list_next(request, response)
            if request:
                time.sleep(API_DELAY_BETWEEN_PAGES)

        return comments

    def fetch_all_replies(
        self,
        parent_id: str,
        video_id: str,
        metadata: Dict[str, Any],
        max_results: Optional[int],
        min_likes: int,
        exclude_creator: bool,
        date_from: Optional[str],
        date_to: Optional[str],
        filter_words: Optional[List[str]],
        cancel_event: Any,
    ) -> List[Dict[str, Any]]:
        replies: List[Dict[str, Any]] = []
        request = self.youtube.comments().list(
            part="snippet",
            parentId=parent_id,
            maxResults=YOUTUBE_REPLIES_PER_PAGE,
            textFormat="plainText",
        )

        while request and not self._cancelled(cancel_event):
            try:
                response = request.execute()
            except Exception:
                # Some reply threads can be private, deleted, or otherwise inaccessible.
                break

            for reply_item in response.get("items", []):
                if self._cancelled(cancel_event) or self._limit_reached(replies, max_results):
                    break

                reply = self._build_comment_row(
                    video_id=video_id,
                    metadata=metadata,
                    comment_id=reply_item.get("id", ""),
                    snippet=reply_item.get("snippet", {}),
                    item_type="Reply",
                    parent_id=parent_id,
                    reply_count=0,
                )

                if self._passes_fetch_filters(
                    reply, min_likes, exclude_creator, date_from, date_to, filter_words, metadata
                ):
                    replies.append(reply)

            request = None if self._limit_reached(replies, max_results) else self.youtube.comments().list_next(request, response)
            if request:
                time.sleep(API_DELAY_BETWEEN_PAGES)

        return replies

    def fetch_live_chat_messages(
        self,
        metadata: Dict[str, Any],
        max_results: Optional[int],
        date_from: Optional[str],
        date_to: Optional[str],
        filter_words: Optional[List[str]],
        cancel_event: Any,
    ) -> List[Dict[str, Any]]:
        live_chat_id = metadata.get("active_live_chat_id")
        if not live_chat_id:
            return []

        messages: List[Dict[str, Any]] = []
        request = self.youtube.liveChatMessages().list(
            liveChatId=live_chat_id,
            part="snippet,authorDetails",
            maxResults=YOUTUBE_LIVE_CHAT_PER_PAGE,
        )

        # Live chats are polling-based; this prevents an active stream from looping forever.
        pages_read = 0
        max_pages = 1 if max_results is None else 1000

        while request and pages_read < max_pages and not self._cancelled(cancel_event):
            try:
                response = request.execute()
            except Exception as e:
                self._raise_friendly_error(e)

            pages_read += 1
            for item in response.get("items", []):
                if self._cancelled(cancel_event) or self._limit_reached(messages, max_results):
                    break

                snippet = item.get("snippet", {})
                author = item.get("authorDetails", {})
                row = {
                    "video_id": metadata.get("video_id", ""),
                    "video_title": metadata.get("title", ""),
                    "channel_title": metadata.get("channel_title", ""),
                    "id": item.get("id", ""),
                    "author": author.get("displayName", "Unknown"),
                    "author_channel_id": author.get("channelId", ""),
                    "text": snippet.get("displayMessage", ""),
                    "likes": 0,
                    "published_at": snippet.get("publishedAt", ""),
                    "updated_at": snippet.get("publishedAt", ""),
                    "type": "Live Chat",
                    "parent_id": "",
                    "reply_count": 0,
                    "spam_reason": "",
                }

                if self._within_date_range(row["published_at"], date_from, date_to) and self._matches_words(row["text"], filter_words):
                    messages.append(row)

            if self._limit_reached(messages, max_results):
                break
            request = self.youtube.liveChatMessages().list_next(request, response)
            if request and max_results is not None:
                time.sleep(API_DELAY_BETWEEN_PAGES)
            else:
                break

        return messages

    def _build_comment_row(
        self,
        video_id: str,
        metadata: Dict[str, Any],
        comment_id: str,
        snippet: Dict[str, Any],
        item_type: str,
        parent_id: str,
        reply_count: int,
    ) -> Dict[str, Any]:
        author_channel = snippet.get("authorChannelId", {}) or {}
        if isinstance(author_channel, dict):
            author_channel_id = author_channel.get("value", "")
        else:
            author_channel_id = str(author_channel)

        return {
            "video_id": video_id,
            "video_title": metadata.get("title", ""),
            "channel_title": metadata.get("channel_title", ""),
            "id": comment_id,
            "author": snippet.get("authorDisplayName", "Unknown"),
            "author_channel_id": author_channel_id,
            "text": snippet.get("textDisplay", ""),
            "likes": int(snippet.get("likeCount", 0) or 0),
            "published_at": snippet.get("publishedAt", ""),
            "updated_at": snippet.get("updatedAt", ""),
            "type": item_type,
            "parent_id": parent_id,
            "reply_count": reply_count,
            "spam_reason": "",
        }

    @staticmethod
    def api_comment_order(sort_by: str) -> str:
        sort_key = str(sort_by or "").lower()
        if "date" in sort_key or "newest" in sort_key or "oldest" in sort_key or "time" in sort_key:
            return "time"
        return "relevance"

    def sort_items(self, items: List[Dict[str, Any]], sort_by: str) -> None:
        sort_key = str(sort_by or "").lower()
        if "oldest" in sort_key:
            items.sort(key=lambda c: c.get("published_at", ""))
        elif "date" in sort_key or "newest" in sort_key or "time" in sort_key:
            items.sort(key=lambda c: c.get("published_at", ""), reverse=True)
        elif "like" in sort_key or "relevance" in sort_key:
            items.sort(key=lambda c: int(c.get("likes", 0) or 0), reverse=True)

    def _passes_fetch_filters(
        self,
        row: Dict[str, Any],
        min_likes: int,
        exclude_creator: bool,
        date_from: Optional[str],
        date_to: Optional[str],
        filter_words: Optional[List[str]],
        metadata: Dict[str, Any],
    ) -> bool:
        if int(row.get("likes", 0) or 0) < int(min_likes or 0):
            return False
        if exclude_creator and row.get("author_channel_id") and row.get("author_channel_id") == metadata.get("channel_id"):
            return False
        if not self._within_date_range(row.get("published_at", ""), date_from, date_to):
            return False
        if not self._matches_words(row.get("text", ""), filter_words):
            return False
        return True

    @staticmethod
    def _within_date_range(published_at: str, date_from: Optional[str], date_to: Optional[str]) -> bool:
        if not published_at:
            return True
        date_value = published_at[:10]
        if date_from and date_value < date_from:
            return False
        if date_to and date_value > date_to:
            return False
        return True

    @staticmethod
    def _matches_words(text: str, filter_words: Optional[List[str]]) -> bool:
        if not filter_words:
            return True
        lowered = (text or "").lower()
        return any(word.lower() in lowered for word in filter_words if word)

    @staticmethod
    def _limit_reached(items: List[Dict[str, Any]], max_results: Optional[int]) -> bool:
        return max_results is not None and max_results > 0 and len(items) >= max_results

    def _cancelled(self, cancel_event: Any) -> bool:
        return self._should_stop or bool(cancel_event and cancel_event.is_set())

    def is_spam(self, text: str) -> Tuple[bool, str]:
        text = text or ""
        if not text.strip():
            return True, "Empty Comment"
        if re.search(r"(?:\+?\d{1,3}[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}", text):
            return True, "Contains Phone Number"
        if re.search(r"https?://\S+|www\.\S+", text):
            return True, "Contains External Link"
        return False, ""

    def apply_local_filters(
        self,
        items: List[Dict[str, Any]],
        filter_spam: bool = True,
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        clean_items: List[Dict[str, Any]] = []
        spam_items: List[Dict[str, Any]] = []

        for item in items:
            text = item.get("text", "") or ""
            lowered = text.lower()

            is_whitelisted = any(pattern.lower() in lowered for pattern in self.whitelist_patterns)
            is_blacklisted = any(pattern.lower() in lowered for pattern in self.blacklist_patterns)
            is_spam, reason = self.is_spam(text)

            if is_whitelisted:
                item["spam_reason"] = ""
                clean_items.append(item)
            elif is_blacklisted:
                item["spam_reason"] = "Custom Blacklist"
                spam_items.append(item)
            elif filter_spam and is_spam and self.spam_threshold > 0:
                item["spam_reason"] = reason
                spam_items.append(item)
            else:
                item["spam_reason"] = ""
                clean_items.append(item)

        return clean_items, spam_items

    @staticmethod
    def _fieldnames(rows: Iterable[Dict[str, Any]]) -> List[str]:
        preferred = [
            "video_id", "video_title", "channel_title", "id", "author", "author_channel_id",
            "text", "likes", "published_at", "updated_at", "type", "parent_id",
            "reply_count", "spam_reason",
        ]
        keys = []
        for row in rows:
            for key in row.keys():
                if key not in keys:
                    keys.append(key)
        return [key for key in preferred if key in keys] + [key for key in keys if key not in preferred]

    def save_to_csv(
        self,
        metadata_list: List[Dict[str, Any]],
        comments: List[Dict[str, Any]],
        base_filename: str,
        spam_list: Optional[List[Dict[str, Any]]] = None,
    ) -> None:
        self._write_csv(f"{base_filename}_metadata.csv", metadata_list)
        self._write_csv(f"{base_filename}_comments.csv", comments)
        if spam_list:
            self._write_csv(f"{base_filename}_spam.csv", spam_list)

    def _write_csv(self, filename: str, rows: List[Dict[str, Any]]) -> None:
        fieldnames = self._fieldnames(rows) if rows else []
        with open(filename, "w", newline="", encoding="utf-8-sig") as f:
            if not fieldnames:
                return
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)

    def save_to_excel(
        self,
        metadata_list: List[Dict[str, Any]],
        comments: List[Dict[str, Any]],
        filename: str,
        spam_list: Optional[List[Dict[str, Any]]] = None,
    ) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Metadata"
        self._write_sheet(ws, metadata_list)

        ws_comments = wb.create_sheet("Comments")
        self._write_sheet(ws_comments, comments)

        if spam_list:
            ws_spam = wb.create_sheet("Flagged Spam")
            self._write_sheet(ws_spam, spam_list)

        wb.save(filename)

    def _write_sheet(self, ws: Any, rows: List[Dict[str, Any]]) -> None:
        if not rows:
            ws.append(["No data"])
            return
        fieldnames = self._fieldnames(rows)
        ws.append(fieldnames)
        for row in rows:
            ws.append([row.get(field, "") for field in fieldnames])

    def save_to_txt(
        self,
        metadata_list: List[Dict[str, Any]],
        comments: List[Dict[str, Any]],
        filename: str,
        spam_list: Optional[List[Dict[str, Any]]] = None,
    ) -> None:
        """Save comments/replies in a Notepad-friendly reading format."""
        def safe(value: Any) -> str:
            return "" if value is None else str(value)

        def write_item(f, row: Dict[str, Any], number: Optional[int] = None, indent: str = "") -> None:
            prefix = f"[{number}] " if number is not None else ""
            item_type = safe(row.get("type", "Comment"))
            f.write(f"{indent}{prefix}{item_type}\n")
            f.write(f"{indent}Author: {safe(row.get('author'))}\n")
            f.write(f"{indent}Date: {safe(row.get('published_at'))}\n")
            f.write(f"{indent}Likes: {safe(row.get('likes', 0))}\n")
            if row.get("parent_id"):
                f.write(f"{indent}Parent ID: {safe(row.get('parent_id'))}\n")
            if row.get("reply_count") not in (None, "", 0, "0"):
                f.write(f"{indent}Reported replies: {safe(row.get('reply_count'))}\n")
            if row.get("spam_reason"):
                f.write(f"{indent}Spam reason: {safe(row.get('spam_reason'))}\n")
            f.write(f"{indent}Text:\n")
            text_lines = safe(row.get("text", "")).splitlines() or [""]
            for line in text_lines:
                f.write(f"{indent}  {line}\n")
            f.write("\n")

        replies_by_parent: Dict[str, List[Dict[str, Any]]] = {}
        parents: List[Dict[str, Any]] = []
        live_chat: List[Dict[str, Any]] = []
        orphan_replies: List[Dict[str, Any]] = []

        for row in comments:
            row_type = safe(row.get("type", "")).lower()
            if row_type == "reply":
                parent_id = safe(row.get("parent_id"))
                if parent_id:
                    replies_by_parent.setdefault(parent_id, []).append(row)
                else:
                    orphan_replies.append(row)
            elif row_type == "live chat":
                live_chat.append(row)
            else:
                parents.append(row)

        parent_ids = {safe(row.get("id")) for row in parents}
        for parent_id, rows in replies_by_parent.items():
            if parent_id not in parent_ids:
                orphan_replies.extend(rows)

        with open(filename, "w", encoding="utf-8") as f:
            f.write("YouTube Comment Extractor - Readable Export\n")
            f.write("=" * 70 + "\n\n")

            if metadata_list:
                for idx, meta in enumerate(metadata_list, start=1):
                    f.write(f"Video {idx}: {safe(meta.get('title'))}\n")
                    f.write(f"Channel: {safe(meta.get('channel_title'))}\n")
                    f.write(f"Video ID: {safe(meta.get('video_id'))}\n")
                    f.write(f"Public comment count: {safe(meta.get('comment_count'))}\n")
                    f.write(f"Items fetched: {safe(meta.get('items_fetched'))}\n")
                    f.write("-" * 70 + "\n")
                f.write("\n")

            f.write(f"Clean exported items: {len(comments)}\n")
            f.write(f"Flagged spam items: {len(spam_list or [])}\n")
            f.write("\n" + "=" * 70 + "\n")
            f.write("COMMENTS AND REPLIES\n")
            f.write("=" * 70 + "\n\n")

            for number, parent in enumerate(parents, start=1):
                write_item(f, parent, number=number)
                for reply in replies_by_parent.get(safe(parent.get("id")), []):
                    f.write("    ↳ Reply\n")
                    write_item(f, reply, indent="    ")
                f.write("-" * 70 + "\n\n")

            if orphan_replies:
                f.write("ORPHAN REPLIES / REPLIES WHOSE PARENT WAS FILTERED OUT\n")
                f.write("=" * 70 + "\n\n")
                for reply in orphan_replies:
                    write_item(f, reply)
                    f.write("-" * 70 + "\n\n")

            if live_chat:
                f.write("LIVE CHAT\n")
                f.write("=" * 70 + "\n\n")
                for number, row in enumerate(live_chat, start=1):
                    write_item(f, row, number=number)
                    f.write("-" * 70 + "\n\n")

            if spam_list:
                f.write("FLAGGED SPAM\n")
                f.write("=" * 70 + "\n\n")
                for number, row in enumerate(spam_list, start=1):
                    write_item(f, row, number=number)
                    f.write("-" * 70 + "\n\n")

    def _raise_friendly_error(self, error: Exception) -> None:
        error_text = str(error).lower()
        if "quota" in error_text:
            raise QuotaExceededError("API quota exceeded.") from error
        if "commentsdisabled" in error_text or "comments disabled" in error_text or "disabled" in error_text:
            raise CommentsDisabledError("Comments are disabled for this video.") from error
        if "notfound" in error_text or "video not found" in error_text:
            raise VideoNotFoundError("The requested video could not be found.") from error
        raise RuntimeError(f"API Error: {error}") from error
